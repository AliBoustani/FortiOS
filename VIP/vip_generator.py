import sys
import getpass

try:
    from openpyxl import load_workbook

except ImportError:
    print " ' Openpyxl ' package for Python is not installed on your system.\n"
    print "Get it from https://pypi.org/project/openpyxl/  and try again."
    sys.exit()


try:
    from netmiko import ConnectHandler

except ImportError:
    print " ' Netmiko ' package for Python is not installed on your system.\n"
    print "Get it from https://pypi.org/project/netmiko/  and try again."
    sys.exit()



try:
    #Specify your excel document! This is my test sample
    wb = load_workbook(filename='MyDoc.xlsx')

    #The sheet name in your excel file
    ws = wb['Sheet1']

    #Determining the row & column size
    maxrow = ws.max_row
    maxcols = ws.max_column

    #Enter your FortiGate IP address and credentials
    host = raw_input("Please Enter your FortiGate device IP address: ")
    username = raw_input("Please Enter your FortiGate username: ")
    password = getpass.getpass()


    #Creating connection profile for your device
    My_fortigate = {
        'device_type': 'fortinet',
        'ip': host,
        'username': username,
        'password': password,
        }

    #The function!
    def fg_vip_conf(vip_name, intip, extip, port):

        #VIP CLI script
        commands_vip = ['config firewall vip', 'edit '+vip_name+'\-Port\ '+port,
'set extip '+ extip ,'set extintf any','set mappedip '+ intip,'set portforward enable','set protocol tcp','set extport '+port,'set mappedport '+port, 'end']
    
        #VIP Group CLI script
        commands_vipgrp = ['config firewall vipgrp', 'edit '+ vip_name,'set interface any', 'append member '+vip_name+'\-Port\ '+port, 'end']

        output_vip = net_connect.send_config_set(commands_vip)
        output_vipgrp = net_connect.send_config_set(commands_vipgrp)

       # print(output_vip)
       # print(output_vipgrp)





    #Establishing SSH connection
    net_connect = ConnectHandler(**My_fortigate)


    #Applying th function for each column in each row
    #It starts from second raw! First raw is just about each column informations.
    for j in range (2, maxrow+1):
        dic1 = dict((i, str(ws.cell(row=j, column=i).value)) for i in range(1, maxcols+1))
        for k in range(4, maxcols+1):  #Considering the specified ports start at column 4 !
            if dic1[k] != 'None':
                fg_vip_conf(dic1[1], dic1[2], dic1[3], dic1[k])
            print("\n *** VIP & VIP Group have created for: " +dic1[1])

        break


    #Terminate the SSH connection
    net_connect.disconnect() 

except KeyboardInterrupt:
    print "\n\nProgram aborted by user. Exiting...\n"
    sys.exit()            

#End of program
